import streamlit as st
import pandas as pd
import openpyxl
import io
import zipfile
from datetime import datetime

def main():
    st.set_page_config(page_title="명세서 완벽 제어 자동화", layout="wide")
    
    st.title("🎯 원본 엑셀 자동 수정 및 빈 줄 삭제기")
    st.write("첫 번째 시트 서식을 100% 보존하며, 정확한 로직으로 값만 수정하고 남는 줄을 삭제합니다.")

    uploaded_file = st.file_uploader("명세서 엑셀 파일을 업로드하세요 (.xlsx)", type=['xlsx'])

    if uploaded_file:
        try:
            # 두 번째 시트 로드 (수량 데이터)
            xls = pd.ExcelFile(uploaded_file)
            df_matrix = pd.read_excel(xls, sheet_name=xls.sheet_names[1], header=2)
            
            if st.button("명세서 엑셀 일괄 생성"):
                zip_buffer = io.BytesIO()
                
                with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_f:
                    # 지점명 추출
                    centers = [c for c in df_matrix.columns if c not in ['제품명', '규격', 'Total'] and not pd.isna(c)]
                    
                    for center in centers:
                        is_bonus = '할증' in str(center) or '힐증' in str(center)
                        
                        # ★ 도화지가 될 원본 엑셀 로드
                        uploaded_file.seek(0) 
                        wb = openpyxl.load_workbook(uploaded_file, data_only=False)
                        ws = wb.worksheets[0] 
                        
                        # ----------------------------------------------------
                        # 1. 엑셀 내에서 "제품명", "낱개가격" 등의 열(Column) 위치 자동 추적
                        # ----------------------------------------------------
                        header_row = -1
                        col_name = col_qty = col_price = col_total = col_inbox = col_box = -1
                        
                        for r in range(1, 30):
                            for c in range(1, 20):
                                val = str(ws.cell(row=r, column=c).value).replace(" ", "")
                                if "제품명" in val:
                                    header_row = r; col_name = c
                                elif "낱개수" in val or "날개수" in val: col_qty = c
                                elif "낱개가격" in val or "날개가격" in val or "단가" in val: col_price = c
                                elif "공급가액" in val: col_total = c
                                elif "입수" in val: col_inbox = c
                                elif "Box" in val or "box" in val.lower(): col_box = c
                            if header_row != -1:
                                break
                        
                        # ----------------------------------------------------
                        # 2. "배송처:" 문구 찾아서 지점명으로 변경
                        # ----------------------------------------------------
                        for r in range(1, header_row):
                            for c in range(1, 20):
                                val = str(ws.cell(row=r, column=c).value)
                                if "배송처" in val:
                                    ws.cell(row=r, column=c).value = f"배송처: {str(center).strip()}"
                        
                        # ----------------------------------------------------
                        # 3. 데이터 끝나는 지점 ("합계") 확인
                        # ----------------------------------------------------
                        footer_row = -1
                        for r in range(header_row + 1, 200):
                            for c in range(1, 20):
                                val = str(ws.cell(row=r, column=c).value).replace(" ", "")
                                if "합계" in val:
                                    footer_row = r
                                    break
                            if footer_row != -1:
                                break
                        
                        # 위치를 찾지 못했으면 패스
                        if header_row == -1 or footer_row == -1:
                            continue
                            
                        sum_total = 0
                        rows_kept = 0
                        
                        # ----------------------------------------------------
                        # 4. 행 삭제가 꼬이지 않게 '아래에서 위로(역순)' 돌면서 작업
                        # ----------------------------------------------------
                        for r in range(footer_row - 1, header_row, -1):
                            p_name = ws.cell(row=r, column=col_name).value
                            
                            # 비어있는 행이거나 불필요한 행이면 삭제
                            if not p_name or str(p_name).strip() == "":
                                ws.delete_rows(r, 1)
                                continue
                                
                            p_name_clean = str(p_name).strip()
                            
                            # 두 번째 시트에서 낱개수 조회
                            match = df_matrix[df_matrix['제품명'].astype(str).str.strip() == p_name_clean]
                            qty = 0
                            if not match.empty:
                                val = match.iloc[0][center]
                                if pd.notna(val) and str(val).replace('.0','').isdigit():
                                    qty = int(float(val))
                                    
                            # [조건 1] 해당 지점에 배송할 제품이 있는 경우
                            if qty > 0:
                                rows_kept += 1
                                ws.cell(row=r, column=col_qty).value = qty # 낱개수 입력
                                
                                # [조건 2] 원본 시트에 적힌 낱개가격 읽어오기
                                price_val = ws.cell(row=r, column=col_price).value
                                try:
                                    price = int(float(str(price_val).replace(',', '')))
                                except:
                                    price = 0
                                    
                                # 할증 지점인 경우만 가격을 0원으로 덮어씀
                                if is_bonus:
                                    price = 0
                                    ws.cell(row=r, column=col_price).value = 0
                                    ws.cell(row=r, column=col_name).value = f"{p_name_clean} (할증분)"
                                    
                                # [조건 3] 수량 × 가격 = 공급가액 계산 및 입력
                                total = qty * price
                                ws.cell(row=r, column=col_total).value = total
                                sum_total += total
                                
                                # (선택) Box 수 계산
                                inbox_val = ws.cell(row=r, column=col_inbox).value
                                try:
                                    inbox = int(float(str(inbox_val)))
                                    if inbox > 0:
                                        ws.cell(row=r, column=col_box).value = qty // inbox
                                except:
                                    pass
                                    
                            # [조건 4] 해당 안 되는 제품은 행 전체를 완벽히 삭제
                            else:
                                ws.delete_rows(r, 1)
                                
                        # ----------------------------------------------------
                        # 5. 행이 삭제되어 위로 딸려 올라간 "합계"를 다시 찾아 표시
                        # ----------------------------------------------------
                        if rows_kept > 0:
                            for r in range(header_row + 1, 200):
                                for c in range(1, 20):
                                    val = str(ws.cell(row=r, column=c).value).replace(" ", "")
                                    if "합계" in val:
                                        # 찾은 합계 줄의 '공급가액(total)' 열에 총액 입력
                                        ws.cell(row=r, column=col_total).value = sum_total
                                        break
                                        
                            # 엑셀 파일 저장 및 ZIP에 추가
                            excel_io = io.BytesIO()
                            wb.save(excel_io)
                            excel_io.seek(0)
                            
                            safe_center = str(center).replace('\n', ' ').replace('/', '_').strip()
                            zip_f.writestr(f"거래명세서_{safe_center}.xlsx", excel_io.read())
                
                # 결과물 다운로드
                st.download_button(
                    label="📦 요청사항 100% 반영된 엑셀 일괄 다운로드 (ZIP)", 
                    data=zip_buffer.getvalue(), 
                    file_name=f"거래명세서_최종본_{datetime.now().strftime('%m%d')}.zip",
                    mime="application/zip"
                )
                st.success("데이터 수정, 빈칸 줄 삭제, 합계 및 배송처 변경이 모두 완료되었습니다!")

        except Exception as e:
            st.error(f"오류가 발생했습니다: {e}")

if __name__ == "__main__":
    main()
